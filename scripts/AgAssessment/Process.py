# --------------------------------------------------------------------------------
# Name:        Process
# Purpose:     This tool performs all of the ag assessment calculations
#
# License:     GNU Affero General Public License v3.
#              Full license in LICENSE file, or at <https://www.gnu.org/licenses/>
# --------------------------------------------------------------------------------

import csv
import json
import arcpy
import pathlib
import openpyxl

from ..helpers import sanitize, license, toggle_required_parameter, reload_module
from ..helpers import print_messages as log
from ..helpers import setup_environment as setup
from ..helpers import validate_spatial_reference as validate

class Process(object):
    def __init__(self):
        """Define the tool (tool name is the name of the class)."""
        self.label = "3. Process"
        self.category = "Automated Ag Assessment"
        self.description = "Run after splitting parcels into use areas"

    def getParameterInfo(self):
        """Define parameter definitions"""
        param0 = arcpy.Parameter(
            displayName="Soils",
            name="soils",
            datatype="GPFeatureLayer",
            parameterType="Required",
            direction="Input")
        param0.filter.list = ["Polygon"]

        param1 = arcpy.Parameter(
            displayName="Soils MUSYM Field",
            name="soils_musym_field",
            datatype="GPString",
            parameterType="Required",
            direction="Input")
        param1.filter.type = "ValueList"
        param1.filter.list = []

        param2 = arcpy.Parameter(
            displayName="Soils MUKEY Field",
            name="soils_mukey_field",
            datatype="GPString",
            parameterType="Required",
            direction="Input")
        param2.filter.type = "ValueList"
        param2.filter.list = []

        params = [param0, param1, param2]
        return params

    def updateParameters(self, parameters):
        # get soils MUSYM nad MUKEY field
        if not parameters[0].hasBeenValidated:
            if parameters[0].value:
                fields = [f.name for f in arcpy.ListFields(parameters[0].value)]
                parameters[1].enabled = True
                parameters[2].enabled = True
                parameters[1].filter.list = fields
                parameters[2].filter.list = fields
                if "MUSYM" in fields:
                    parameters[1].value = "MUSYM"
                if "MUKEY" in fields:
                    parameters[2].value = "MUKEY"
            else:
                parameters[1].enabled = False
                parameters[2].enabled = False

        return

    def updateMessages(self, parameters):
        """Modify the messages created by internal validation for each tool parameter."""
        # make newly toggled on parameters required
        toggle_required_parameter(parameters[0], parameters[1])
        toggle_required_parameter(parameters[0], parameters[2])

        validate(parameters)
        return

    def isLicensed(self):
        """Set whether the tool is licensed to execute."""
        return license()

    @reload_module(__name__)
    def execute(self, parameters, messages):
        """The source code of the tool."""
        # Setup
        log("setting up project")
        project, active_map = setup()
        project_dir = project.homeFolder
        cache_file_path = "{}/.ag_cache.json".format(project_dir)

        # read in json
        log("reading in cache")
        cache = {}
        with open(cache_file_path) as file:
            cache = json.load(file)
        parcels = cache["parcels"]
        output_folder = cache["output_folder"]

        # Parameters
        log("reading in parameters")
        soil_layer = parameters[0].value
        soils_musym = parameters[1].value
        soils_mukey = parameters[2].value

        # collect layouts to be able to close and redisplay later
        layouts = []
        log("iterating through parcels and processing")
        for parcel in parcels:
            # find map of parcel
            m = None
            try:
                m = project.listMaps(parcel)[0]
            except:
                log("unable to find map for {}, results may be incomplete".format(parcel))
                continue

            # Clear selection
            m.clearSelection()

            # find layout
            lyt = None
            try:
                lyt = project.listLayouts(parcel)[0]
                layouts.append(lyt)
            except:
                log("couldn't find layout for parcel {}, results may be incomplete".format(parcel))
                continue

            # Helper variables
            soils_layers = []
            use_layers = []
            tables = []

            # Start work
            parcel_last_four = sanitize(parcel)[-4:]
            lyrs = m.listLayers("*_{}".format(parcel_last_four))
            lyr_types = set()
            log("processing {}".format(parcel))
            for lyr in lyrs:
                # Update symbology
                lyr_type = ""
                if "agland" in lyr.name.lower():
                    use_layers.append(lyr)
                    lyr_type = "Agland"
                elif "nonag" in lyr.name.lower():
                    use_layers.append(lyr)
                    lyr_type = "NonAg"
                elif "forest" in lyr.name.lower():
                    use_layers.append(lyr)
                    lyr_type = "Forest"
                else:
                    continue
                lyr_types.add(lyr_type)

                # Create clip layer
                new_layer_path = "{}\\{}".format(arcpy.env.workspace, "{}_{}_soils".format(sanitize(lyr.name), sanitize(parcel)))
                arcpy.analysis.Clip(soil_layer, lyr, new_layer_path)

                # Dissolve duplicate MUSYMs
                dissolve_layer_path = "{}\\{}".format(arcpy.env.workspace, "{}_{}_soils_dissolved".format(sanitize(lyr.name), sanitize(parcel)))
                arcpy.management.Dissolve(new_layer_path, dissolve_layer_path, [soils_musym,soils_mukey])

                # Add to map
                new_layer = m.addDataFromPath(dissolve_layer_path)
                soils_layers.append(new_layer)
                new_layer.name = "{}_soils".format(lyr.name)

                # Add acreage field
                field_alias = "{} Acres".format(lyr_type)
                arcpy.management.AddField(new_layer, "Acres", "FLOAT", 2, 2, field_alias=field_alias)

                # Calculate geometry
                arcpy.management.CalculateGeometryAttributes(in_features=new_layer.name, geometry_property=[["Acres", "AREA_GEODESIC"]], area_unit="ACRES_US")

                # Update soils clip layer symbology
                sym = new_layer.symbology
                sym.renderer.symbol.color = {'RGB' : [0, 0, 0, 0]}
                sym.renderer.symbol.outlineColor = {'RGB' : [255, 255, 0, 100]}
                sym.renderer.symbol.size = 1.5
                new_layer.symbology = sym

                # Add label
                new_layer.showLabels = True
                label_class = new_layer.listLabelClasses()[0]
                label_class.visible = True
                label_class.expression = "$feature.{}".format(soils_musym)

                l_cim = new_layer.getDefinition('V3')
                lc = l_cim.labelClasses[0]

                # Update text properties of label
                lc.textSymbol.symbol.height = 12
                lc.textSymbol.symbol.symbol.symbolLayers = [
                    {
                        "type": "CIMSolidFill",
                        "enable": True,
                        "color": {
                            "type": "CIMRGBColor",
                            "values": [255, 255, 0, 100]
                        }
                    }
                ]
                lc.standardLabelPlacementProperties.numLabelsOption = "OneLabelPerPart"

                # Update CIM definition
                new_layer.setDefinition(l_cim)

                # Get soils layer attribute table and export / extract needed fields for layout
                table_path = "{}\\{}".format(arcpy.env.workspace, "{}_ExportTable".format(sanitize(new_layer.name)))
                arcpy.conversion.ExportTable(new_layer.name, table_path)
                arcpy.management.DeleteField(table_path, ["{}".format(soils_musym), "Acres", "{}".format(soils_mukey)], "KEEP_FIELDS")

                # Add soils table export to the given map
                soils_table = arcpy.mp.Table(table_path)
                tables.append(soils_table)
                m.addTable(soils_table)
                soils_table_uri = soils_table.URI

                # Get layout table
                tbl = lyt.listElements("MAPSURROUND_ELEMENT", lyr_type)[0]

                # Set layout table to exported attributes table
                tbl_cim = tbl.getDefinition("V3")
                tbl_cim.mapMemberURI = soils_table_uri
                tbl.setDefinition(tbl_cim)

                # Refresh layout
                lyt_cim = lyt.getDefinition('V3')
                lyt.setDefinition(lyt_cim)

                project.save()
                project.closeViews("LAYOUTS")

            # Reorder layers so soils layers are last
            log("reordering layers for {}".format(parcel))
            for soils_layer in soils_layers:
                for use_layer in use_layers:
                    m.moveLayer(use_layer, soils_layer, "AFTER")

            # Remove unused tables
            log("removing unused tables for {}".format(parcel))
            uses = {'Agland', 'Forest', 'NonAg'}
            for i in uses:
                if i not in lyr_types:
                    tbl_remove = lyt.listElements("MAPSURROUND_ELEMENT", i)[0]
                    lyt.deleteElement(tbl_remove)

            # Display wanted legend items only
            log("removing unused legend items for {}".format(parcel))
            legend = lyt.listElements("LEGEND_ELEMENT")[0]
            legend_items = legend.items
            use_layer_names = [ i.name for i in use_layers ]
            for item in legend_items:
                if item.name in use_layer_names:
                    item.visible = True
                else:
                    item.visible = False

            # Export tables
            log("exporting tables for {}".format(parcel))
            soils_tables = []
            for table in tables:
                table_file_path = "{}\{}.csv".format(output_folder, table.name)
                soils_tables.append(table_file_path)
                arcpy.conversion.ExportTable(table, table_file_path)

            # Soil group worksheet
            sgw_path = "{}\{}.xlsx".format(output_folder, lyt.name)

            # Populate soil group worksheet with values
            log("filling out soil group worksheet for {}".format(parcel))
            sgw_path = pathlib.PureWindowsPath(sgw_path).as_posix()
            sgw_workbook = openpyxl.load_workbook(sgw_path)
            ws = sgw_workbook['SGW']
            for tbl in soils_tables:
                if "agland" in tbl.lower():
                    with open(tbl, 'r') as csvfile:
                        csvreader = csv.reader(csvfile)
                        next(csvreader)
                        idx = 0
                        for row in csvreader:
                            if idx < 24:
                                soil_cell = 'A{}'.format(34 + idx)
                                area_cell = 'H{}'.format(34 + idx)
                                mukey_cell = 'F{}'.format(34 + idx)
                                ws[soil_cell] = row[0]
                                ws[mukey_cell] = int(row[1])
                                ws[area_cell] = round(float(row[2]), 2)
                            else:
                                # overflow
                                soil_cell = 'N{}'.format(9 + idx)
                                area_cell = 'U{}'.format(9 + idx)
                                mukey_cell = 'S{}'.format(9 + idx)
                                ws[soil_cell] = row[0]
                                ws[mukey_cell] = int(row[1])
                                ws[area_cell] = round(float(row[2]), 2)
                            idx += 1
                elif "nonag" in tbl.lower():
                    tot = 0
                    with open(tbl, 'r') as csvfile:
                        csvreader = csv.reader(csvfile)
                        next(csvreader)
                        for row in csvreader:
                            tot += float(row[2])
                    ws['K28'] = tot
                elif "forest" in tbl.lower():
                    tot = 0
                    with open(tbl, 'r') as csvfile:
                        csvreader = csv.reader(csvfile)
                        next(csvreader)
                        for row in csvreader:
                            tot += float(row[2])
                    ws['L24'] = tot
            sgw_workbook.save(sgw_path)
            sgw_workbook.close()
            del sgw_workbook
            del ws

        # open layouts
        log("opening layouts")
        for layout in layouts:
            layout.openView()

        # Save
        log("saving project")
        project.save()
        del project

        return
