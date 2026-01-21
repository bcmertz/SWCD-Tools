# --------------------------------------------------------------------------------
# Name:        Delineate
# Purpose:     This tool finds a parcel and sets up an ag assessment project
#
# License:     GNU Affero General Public License v3.
#              Full license in LICENSE file, or at <https://www.gnu.org/licenses/>
# --------------------------------------------------------------------------------

import os
import json
import arcpy
import shutil
import pathlib
import openpyxl

from helpers import license, sanitize, toggle_required_parameter, reload_module
from helpers import print_messages as log
from helpers import setup_environment as setup
from helpers import validate_spatial_reference as validate

class Delineate(object):
    def __init__(self):
        """Define the tool (tool name is the name of the class)."""
        self.label = "1. Delineate Parcels"
        self.description = "Delienate parcels and create folder structure"
        self.category = "Automated Ag Assessment"
        self.canRunInBackground = False

    def getParameterInfo(self):
        """Define parameter definitions"""
        param0 = arcpy.Parameter(
            displayName="Parcels Layer",
            name="parcels",
            datatype="GPFeatureLayer",
            parameterType="Required",
            direction="Input")
        param0.filter.list = ["Polygon"]

        param1 = arcpy.Parameter(
            displayName="Tax Parcel ID Field",
            name="parcel_id_field",
            datatype="GPString",
            parameterType="Optional",
            direction="Input")
        param1.filter.type = "ValueList"
        param1.filter.list = []

        param2 = arcpy.Parameter(
            displayName="SWIS Code Field",
            name="swis_field",
            datatype="GPString",
            parameterType="Optional",
            direction="Input")
        param2.filter.type = "ValueList"
        param2.filter.list = []

        param3 = arcpy.Parameter(
            displayName="Municipality Field",
            name="municipality_field",
            datatype="GPString",
            parameterType="Optional",
            direction="Input")
        param3.filter.type = "ValueList"
        param3.filter.list = []

        param4 = arcpy.Parameter(
            displayName="Address Field",
            name="address_field",
            datatype="GPString",
            parameterType="Optional",
            direction="Input")
        param4.filter.type = "ValueList"
        param4.filter.list = []

        param5 = arcpy.Parameter(
            displayName="Ag District Field",
            name="ag_dist_field",
            datatype="GPString",
            parameterType="Optional",
            direction="Input")
        param5.filter.type = "ValueList"
        param5.filter.list = []

        param6 = arcpy.Parameter(
            displayName="Tax Parcel ID Number",
            name="tax_id_number",
            datatype="GPString",
            parameterType="Required",
            direction="Input",
            multiValue=True)

        param7 = arcpy.Parameter(
            displayName="Last Name",
            name="last_name",
            datatype="GPString",
            parameterType="Required",
            direction="Input")

        param8 = arcpy.Parameter(
            displayName="First Name",
            name="first_name",
            datatype="GPString",
            parameterType="Required",
            direction="Input")

        param9 = arcpy.Parameter(
            displayName="Mailing Street Name and Number",
            name="street_name_num",
            datatype="GPString",
            parameterType="Required",
            direction="Input")

        param10 = arcpy.Parameter(
            displayName="Mailing City/Town",
            name="city_town",
            datatype="GPString",
            parameterType="Required",
            direction="Input")

        param11 = arcpy.Parameter(
            displayName="Mailing State (two letter)",
            name="state",
            datatype="GPString",
            parameterType="Required",
            direction="Input")

        param12 = arcpy.Parameter(
            displayName="Mailing Zip Code",
            name="zip_code",
            datatype="GPString",
            parameterType="Required",
            direction="Input")

        param13 = arcpy.Parameter(
            displayName="Output Folder",
            name="output_location",
            datatype="DEFolder",
            parameterType="Required",
            direction="Input")

        params = [param0, param1, param2, param3, param4, param5, param6, param7, param8, param9, param10, param11, param12, param13]
        return params

    def updateParameters(self, parameters):
        # get Parcel Id, SWIS, Municipality, Address, and Ag District fields
        # set value if default field exists
        if not parameters[0].hasBeenValidated:
            if parameters[0].value:
                fields = [f.name for f in arcpy.ListFields(parameters[0].value)]
                parameters[1].enabled = True
                parameters[2].enabled = True
                parameters[3].enabled = True
                parameters[4].enabled = True
                parameters[5].enabled = True
                parameters[1].filter.list = fields
                parameters[2].filter.list = fields
                parameters[3].filter.list = fields
                parameters[4].filter.list = fields
                parameters[5].filter.list = fields
                if "PRINT_KEY" in fields:
                    parameters[1].value = "PRINT_KEY"
                if "SWIS" in fields:
                    parameters[2].value = "SWIS"
                if "TOWN" in fields:
                    parameters[3].value = "TOWN"
                if "LOCATION" in fields:
                    parameters[4].value = "LOCATION"
                if "AGDIST" in fields:
                    parameters[5].value = "AGDIST"
            else:
                parameters[1].enabled = False
                parameters[2].enabled = False
                parameters[3].enabled = False
                parameters[4].enabled = False
                parameters[5].enabled = False

        return

    def updateMessages(self, parameters):
        """Modify the messages created by internal validation for each tool parameter."""
        # make optional parameters required based off of parameters[0]
        toggle_required_parameter(parameters[0], parameters[1])
        toggle_required_parameter(parameters[0], parameters[2])
        toggle_required_parameter(parameters[0], parameters[3])
        toggle_required_parameter(parameters[0], parameters[4])
        toggle_required_parameter(parameters[0], parameters[5])

        validate(parameters)

        return

    def isLicensed(self):
        """Set whether the tool is licensed to execute."""
        return license()

    @reload_module(__name__)
    def execute(self, parameters, messages):
        """The source code of the tool."""
        # Setup
        log("setting up project")
        project, active_map = setup()
        project_dir = project.homeFolder
        cache_file_path = "{}/.ag_cache.json".format(project_dir)

        # Parameters
        log("reading in parameters")
        parcel_layer = parameters[0].value
        parcel_layer_field = parameters[1].value
        swis_field = parameters[2].value
        municipality_field = parameters[3].value
        address_field = parameters[4].value
        ag_dist_field = parameters[5].value
        tax_id_nums = parameters[6].valueAsText.split(";")
        last_name = parameters[7].valueAsText
        first_name = parameters[8].valueAsText
        street_name_num = parameters[9].valueAsText
        city_town = parameters[10].valueAsText
        state = parameters[11].valueAsText
        zip_code = parameters[12].valueAsText
        output_folder = parameters[13].valueAsText

        # setup cache
        log("setting up cache")
        cache_json = {
            "parcels": tax_id_nums,
            "output_folder": output_folder,
        }
        with open(cache_file_path, "w") as file:
            json.dump(cache_json, file)

        # clear selections from map
        orig_map = active_map
        orig_map.clearSelection()

        # sgw template path
        sgw_template = os.path.join(os.path.dirname(__file__), os.pardir, os.pardir, 'assets', 'Soil Group Worksheet.xlsx')

        # use layout template
        log("finding layout")
        orig_layout = project.importDocument(os.path.join(os.path.dirname(__file__), os.pardir, os.pardir, 'assets', 'agassessment_layout.pagx'))
        layouts = []

        for tax_id_num in tax_id_nums:
            layer_name = "{}_{}".format(last_name, tax_id_num)
            sanitized_name = sanitize(layer_name)
            parcel_path = "{}\\{}".format(arcpy.env.workspace, sanitized_name)

            # create new map and make it active
            log("creating map for {}".format(tax_id_num))
            new_map = project.copyItem(orig_map, tax_id_num)
            new_map.openView()
            cam = project.activeView.camera

            # create a new layout
            log("creating layout for {}".format(tax_id_num))
            new_layout = project.copyItem(orig_layout, tax_id_num)
            layouts.append(new_layout)
            new_layout.openView()

            # set layout's map to new map created
            mf = new_layout.listElements("MAPFRAME_ELEMENT")[0]
            mf.map = new_map
            mf.name = tax_id_num

            # turn off parcel layer
            parcel_layer.visible = False

            # create sql expression to select correct parcel
            sql_expr="{} = '{}'".format(parcel_layer_field, tax_id_num)

            # create parcel layer and add it to the map
            log("adding parcel {}".format(tax_id_num))
            feat = arcpy.management.MakeFeatureLayer(parcel_layer, layer_name, sql_expr)
            arcpy.management.CopyFeatures(feat, parcel_path)
            lyr = new_map.addDataFromPath(parcel_path)
            lyr.name = layer_name

            # update parcel symbology
            log("updating layer symbology {}".format(tax_id_num))
            sym = lyr.symbology
            sym.renderer.symbol.applySymbolFromGallery("Black Outline (2 pts)")
            lyr.symbology = sym

            # Create soil group worksheets for each layout
            log("creating soil group worksheet for {}".format(tax_id_num))
            sgw_path = r'{}\{}.xlsx'.format(output_folder, new_layout.name)
            sgw_path = pathlib.PureWindowsPath(sgw_path).as_posix()
            shutil.copyfile(sgw_template, sgw_path)

            # set SWIS code in layout
            log("finding property values for {}".format(tax_id_num))
            swis_box = new_layout.listElements("TEXT_ELEMENT", "SWIS")[0]
            swis_value = [row[0] for row in arcpy.da.SearchCursor(parcel_path, swis_field)][0]
            swis_box.text = "SWIS: {}".format(swis_value)

            # set name in layout
            name_box = new_layout.listElements("TEXT_ELEMENT", "Name")[0]
            name_box.text = "{}, {}".format(last_name, first_name)

            # set municipality in layout
            municipality_box = new_layout.listElements("TEXT_ELEMENT", "Municipality")[0]
            municipality_value = [row[0] for row in arcpy.da.SearchCursor(parcel_path, municipality_field)][0]
            municipality_box.text = "{}".format(municipality_value)

            # get property address info
            location_value = [row[0] for row in arcpy.da.SearchCursor(parcel_path, address_field)][0]
            agdist_value = [row[0] for row in arcpy.da.SearchCursor(parcel_path, ag_dist_field)][0]
            if agdist_value in ["", None, " "]:
                agdist_value = "__"
            else:
                agdist_value = "x"

            # set SWIS, municipality, tax map identifier, etc in soil group worksheet
            log("writing values to soil group worksheet {}".format(tax_id_num))
            sgw_workbook = openpyxl.load_workbook(sgw_path)
            ws = sgw_workbook['SGW']
            ws['D24'] = swis_value
            ws['D19'] = municipality_value
            ws['D17'] = location_value
            ws['B20'] = agdist_value
            ws['D26'] = tax_id_num
            ws['F13'] = first_name
            ws['B13'] = last_name
            ws['B15'] = street_name_num
            ws['F15'] = city_town
            ws['J15'] = state
            ws['K15'] = zip_code
            sgw_workbook.save(sgw_path)
            sgw_workbook.close()
            del sgw_workbook
            del ws

            # zoom to layer in map object
            log("zooming map to {}".format(tax_id_num))
            ext = arcpy.Describe(lyr).extent
            cam.setExtent(ext)

            # zoom layout to last active map
            mf = new_layout.listElements("MAPFRAME_ELEMENT")[0]
            mf.camera.setExtent(mf.getLayerExtent(lyr))
            mf.camera.scale = mf.camera.scale * 1.1

            # Need to close layouts for camera change to take effect
            project.closeViews("LAYOUTS")

        # export parcel layouts to folder
        log("exporting layouts")
        for layout in layouts:
            layout_file_path = "{}\{}.pdf".format(output_folder, layout.name)
            layout.exportToPDF(layout_file_path)

        # remove unused layout
        project.deleteItem(orig_layout)

        # cleanup
        log("saving project")
        project.save()
        del project

        # open folder to print out maps
        log("opening project folder")
        os.startfile(output_folder)

        return
